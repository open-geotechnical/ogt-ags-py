# -*- coding: utf-8 -*-
"""
@author: Peter Morgan <pete@daffodil.uk.com>
"""
import os
import datetime
from Qt import QtGui, QtCore, Qt, pyqtSignal

from ogt import HAVE_EXCEL
from ogt import ogt_excel
#from img import Ico
import xwidgets
import ogtgui_widgets

if HAVE_EXCEL:
    from openpyxl import load_workbook



class ExcelWorkbookWidget(QtGui.QWidget):

    def __init__(self, parent=None, filename=None):
        QtGui.QWidget.__init__(self, parent)

        self.debug = False
        self.filename = filename
        self.setObjectName("ExcelSheet")

        self.mainLayout = xwidgets.vlayout()
        self.setLayout(self.mainLayout)


        self.tabWidget = QtGui.QTabWidget()
        self.mainLayout.addWidget(self.tabWidget)



    def load_workbook(self, filename=None):



        progress = QtGui.QProgressDialog(self)
        progress.setWindowModality(Qt.WindowModal)
        progress.setLabelText("Loading...\n%s" % filename)
        progress.setRange(0, 0)
        progress.setValue(0)
        progress.forceShow()

        #filename = str(filename)
        print "==", self.filename, type(self.filename), self
        wb = load_workbook(filename=self.filename, data_only=True)
        for sheet in wb.worksheets:
            print sheet
            sheetWidget = ExcelSheetWidget(sheet=sheet)
            self.tabWidget.addTab(sheetWidget, sheet.title)

        progress.hide()



class ExcelSheetWidget(QtGui.QWidget):

    def __init__(self, parent=None, sheet=None):
        QtGui.QWidget.__init__(self, parent)


        self.setObjectName("ExcelSheetWidget")

        self.sheet = sheet

        self.mainLayout = xwidgets.vlayout()
        self.setLayout(self.mainLayout)

        self.table = QtGui.QTableWidget()
        self.mainLayout.addWidget(self.table)

        self.load_sheet()

    def load_sheet(self):
        print "load_sheet", self.sheet, self
        print self.sheet['A1'].value, self.sheet['A2']

        last_col = None

        maxi = 100
        for ridx in range(0, maxi):
            if self.table.rowCount() < ridx + 1:
                self.table.setRowCount(ridx + 1)
            for cidx in range(0, maxi):
                #if self.table.columnCount() < cidx + 1:
                #   self.table.setColumnCount(cidx + 1)
                cell_ref = ogt_excel.rowcol_to_cell(ridx, cidx)
                v = self.sheet[cell_ref].value
                #print ridx, cidx, cell_ref, v
                if v != None and (isinstance(v, str) or isinstance(v, unicode)) and v.upper() == "STOP":
                    last_col = cidx
                    #dsa
                else:
                    if v == None:
                        cv = ""
                    elif isinstance(v, long):
                        cv = str(v)
                    elif isinstance(v, float):
                        cv = str(v)
                    elif isinstance(v, datetime.datetime):
                        cv = str(datetime.datetime)
                    else:
                        cv = v
                    item = QtGui.QTableWidgetItem()
                    item.setText(cv)
                    self.table.setItem(ridx, cidx, item)



class ExpExcelBrowserWidget( QtGui.QWidget ):
    """T"""

    def __init__( self, parent=None):
        QtGui.QWidget.__init__( self, parent )

        self.debug = False
        self.setObjectName("ExpExcelBrowserWidget")

        self.mainLayout = xwidgets.vlayout()
        self.setLayout(self.mainLayout)

        self.splitter = QtGui.QSplitter()
        self.mainLayout.addWidget(self.splitter, 10)

        self.tabWidget = QtGui.QTabWidget()
        self.splitter.addWidget(self.tabWidget)

        self.pathBrowse = ogtgui_widgets.ExpPathBrowseWidget()
        self.splitter.addWidget(self.pathBrowse)

        self.splitter.setStretchFactor(0, 4)
        self.splitter.setStretchFactor(1, 1)


        self.pathBrowse.sigOpenFile.connect(self.on_open_excel)

        self.on_open_excel("/home/ogt/gstl_examples/35579/ATTS/ATTERBERG LIMITS - GSTL 2017 - WH - PAGE 4.xlsx")

    def init_load(self):
        pass


    def on_open_excel(self, filename):
        """
        progress = QtGui.QProgressDialog(self)
        progress.setWindowModality(Qt.WindowModal)
        progress.setLabelText("Loading...\n%s" % filename)
        progress.setRange(0, 0)
        progress.setValue(0)
        progress.forceShow()
        """
        filename = str(filename)

        widget = ExcelWorkbookWidget(filename=filename)
        self.tabWidget.addTab(widget, os.path.basename(filename))
        widget.load_workbook()
        #progress.hide()
